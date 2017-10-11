##############################################################################
#
#    OSIS stands for Open Student Information System. It's an application
#    designed to manage the core business of higher education institutions,
#    such as universities, faculties, institutes and professional schools.
#    The core business involves the administration of students, teachers,
#    courses, programs and so on.
#
#    Copyright (C) 2015-2017 Université catholique de Louvain (http://www.uclouvain.be)
#
#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU General Public License as published by
#    the Free Software Foundation, either version 3 of the License, or
#    (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU General Public License for more details.
#
#    A copy of this license - GNU General Public License - is available
#    at the root of the source code of this program.  If not,
#    see http://www.gnu.org/licenses/.
#
##############################################################################
from django.contrib.auth.decorators import user_passes_test
from openpyxl import load_workbook
from assistant import models as assistant_mdl
from base.views import layout
from base import models as mdl
from django.core.exceptions import ObjectDoesNotExist
from django.contrib import messages
from django.utils.translation import ugettext as _
from assistant.forms import MandateFileForm
from assistant.utils import manager_access
from assistant.models.enums import assistant_type, assistant_phd_inscription, assistant_mandate_renewal
from base.models.enums import entity_type

COLS_NUMBER = 22
ASSISTANTS_IMPORTED = 0
MANDATES_IMPORTED = 0
ASSISTANTS_UPDATED = 0
MANDATES_UPDATED = 0
PERSONS_NOT_FOUND = 0
COLS_TITLES = ['SECTOR', 'FACULTY', 'SCHOOL', 'INSTITUTE', 'POLE', 'SAP_ID', 'GLOBAL_ID', 'LAST_NAME',
               'FIRST_NAME', 'FULLTIME_EQUIVALENT', 'ENTRY_DATE', 'END_DATE', 'ASSISTANT_TYPE_CODE', 'SCALE',
               'CONTRACT_DURATION', 'CONTRACT_DURATION_FTE', 'RENEWAL_TYPE', 'ABSENCES', 'COMMENT', 'OTHER_STATUS',
               'EMAIL', 'FGS']
ASSISTANT_TYPES_ALIASES = {
    'ST': assistant_type.ASSISTANT,
    'AS': assistant_type.TEACHING_ASSISTANT
}


@user_passes_test(manager_access.user_is_manager, login_url='assistants_home')
def upload_mandates_file(request):
    global ASSISTANTS_IMPORTED, ASSISTANTS_UPDATED, MANDATES_IMPORTED, MANDATES_UPDATED, PERSONS_NOT_FOUND
    ASSISTANTS_UPDATED = 0
    ASSISTANTS_IMPORTED = 0
    MANDATES_IMPORTED = 0
    MANDATES_UPDATED = 0
    PERSONS_NOT_FOUND = 0
    if request.method == 'POST':
        form = MandateFileForm(request.POST, request.FILES)
        if form.is_valid():
            file_name = request.FILES['file']
            if file_name is not None:
                try:
                    read_xls_mandates(request, file_name)
                except IndexError:
                    messages.add_message(request, messages.ERROR,
                                         'xls_columns_structure_error'.format('via_excel', 'get_excel_file'))
        else:
            for error_msg in [error_msg for error_msgs in form.errors.values() for error_msg in error_msgs]:
                messages.add_message(request, messages.ERROR, "{}".format(error_msg))
        return show_import_result(request)


@user_passes_test(manager_access.user_is_manager, login_url='assistants_home')
def read_xls_mandates(request, file_name):
    try:
        workbook = load_workbook(file_name, read_only=True, data_only=True)
    except KeyError:
        messages.add_message(request, messages.ERROR, 'file_must_be_xlsx')
        return False
    first_sheet = workbook.get_sheet_names()[0]
    worksheet = workbook.get_sheet_by_name(first_sheet)
    titles_row = []
    current_row = 1
    for row in worksheet.iter_rows():
        if current_row == 1:
            titles_row = save_xls_rows_titles(row)
            if check_file_format(request, titles_row) is False:
                return False
        else:
            current_record = xls_row_to_dict(row, titles_row)
            assistant = create_academic_assistant_if_not_exists(current_record)
            if assistant:
                mandate = create_assistant_mandate_if_not_exists(current_record, assistant)
                sector = search_entity_by_acronym_and_type(current_record.get('SECTOR'),
                                                            entity_type.SECTOR)
                if sector:
                    link_mandate_to_entity(mandate, sector)
                faculty = search_entity_by_acronym_and_type(current_record.get('FACULTY'),
                                                            entity_type.FACULTY)
                if faculty:
                    link_mandate_to_entity(mandate, faculty)
                school = search_entity_by_acronym_and_type(current_record.get('SCHOOL'),
                                                                       entity_type.SCHOOL)
                if school:
                    link_mandate_to_entity(mandate, school)

                institute = search_entity_by_acronym_and_type(current_record.get('INSTITUTE'),
                                                              entity_type.INSTITUTE)
                if institute:
                    link_mandate_to_entity(mandate, institute)

                pole = search_entity_by_acronym_and_type(current_record.get('POLE'), entity_type.POLE)
                if pole:
                    link_mandate_to_entity(mandate, pole)

        current_row += 1


def search_entity_by_acronym_and_type(acronym, type):
    if not acronym:
        return None
    entities = mdl.entity_version.search(entity_type=type, acronym=acronym)
    if len(entities) > 0:
        return entities[0].entity
    else:
        return None


def save_xls_rows_titles(current_row):
    titles = []
    for cell in current_row:
        titles.append(cell.value)
    return titles


def xls_row_to_dict(row, titles):
    record_to_import = {}
    current_col = 0
    for cell in row:
        if titles[current_col] == 'FGS' and len(cell.value) != 8:
            record_to_import[titles[current_col]] = cell.value.zfill(8)
        else:
            record_to_import[titles[current_col]] = cell.value
        current_col += 1
    return record_to_import


def create_academic_assistant_if_not_exists(record):
    global ASSISTANTS_UPDATED, ASSISTANTS_IMPORTED, PERSONS_NOT_FOUND
    person = mdl.person.find_by_global_id(record.get('FGS'))
    if person:
        try:
            assistant = assistant_mdl.academic_assistant.find_by_person(mdl.person.find_by_global_id(record.get('FGS')))
            ASSISTANTS_UPDATED += 1
        except ObjectDoesNotExist:
            assistant = assistant_mdl.academic_assistant.AcademicAssistant()
            ASSISTANTS_IMPORTED += 1
            assistant.person = person
        if record.get('ASSISTANT_TYPE_CODE') == 'AS':
            assistant.inscription = assistant_phd_inscription.NO
        assistant.save()
        return assistant
    else:
        PERSONS_NOT_FOUND += 1
        return None


def create_assistant_mandate_if_not_exists(record, assistant):
    global MANDATES_IMPORTED, MANDATES_UPDATED
    current_academic_year = mdl.academic_year.current_academic_year()
    mandates = assistant_mdl.assistant_mandate.find_mandate(assistant, current_academic_year, record.get('SAP_ID'))
    if len(mandates) == 0:
        mandate = assistant_mdl.assistant_mandate.AssistantMandate()
        mandate.state = 'TO_DO'
        MANDATES_IMPORTED += 1
    else:
        mandate = mandates[0]
        MANDATES_UPDATED += 1
    mandate.assistant = assistant
    mandate.academic_year = current_academic_year
    mandate.end_date = record.get('END_DATE')
    mandate.entry_date = record.get('ENTRY_DATE')
    mandate.fulltime_equivalent = record.get('FULLTIME_EQUIVALENT')
    mandate.sap_id = record.get('SAP_ID')
    mandate.contract_duration = record.get('CONTRACT_DURATION')
    mandate.contract_duration_fte = record.get('CONTRACT_DURATION_FTE')
    if record.get('RENEWAL_TYPE').lower() == 'exceptional' or record.get('RENEWAL_TYPE').lower() == 'exceptionnel':
        mandate.renewal_type = assistant_mandate_renewal.EXCEPTIONAL
    elif record.get('RENEWAL_TYPE').lower() == 'normal':
        mandate.renewal_type = assistant_mandate_renewal.NORMAL
    else:
        mandate.renewal_type = assistant_mandate_renewal.SPECIAL
    mandate.absences = record.get('ABSENCES')
    mandate.comment = record.get('COMMENT')
    mandate.other_status = record.get('OTHER_STATUS')
    if ASSISTANT_TYPES_ALIASES.get(record.get('ASSISTANT_TYPE_CODE')) == assistant_type.TEACHING_ASSISTANT:
        mandate.assistant_type = assistant_type.TEACHING_ASSISTANT
    else:
        mandate.assistant_type = assistant_type.ASSISTANT
    mandate.scale = record.get('SCALE')
    mandate.save()
    return mandate


def link_mandate_to_entity(mandate, entity=None):
    if entity != 'None':
        type = mdl.entity_version.get_by_entity_and_date(entity, None)
        mandate_entities = assistant_mdl.mandate_entity.find_by_mandate_and_type(mandate, type[0].entity_type)
        if len(mandate_entities) > 0:
            mandate_entity = mandate_entities[0]
            mandate_entity.delete()
        mandate_entity = assistant_mdl.mandate_entity.MandateEntity()
        mandate_entity.assistant_mandate = mandate
        mandate_entity.entity = entity
        mandate_entity.save()
        return mandate_entity
    else:
        return None


def show_import_result(request):
    global ASSISTANTS_IMPORTED, ASSISTANTS_UPDATED, MANDATES_IMPORTED, MANDATES_UPDATED, PERSONS_NOT_FOUND
    return layout.render(request, "load_mandates.html", {'imported_assistants': ASSISTANTS_IMPORTED,
                                                         'imported_mandates': MANDATES_IMPORTED,
                                                         'updated_mandates': MANDATES_UPDATED,
                                                         'updated_assistants': ASSISTANTS_UPDATED,
                                                         'persons_not_found': PERSONS_NOT_FOUND})


def check_file_format(request, titles_rows):
    if len(titles_rows) != COLS_NUMBER:
        messages.add_message(request, messages.ERROR, _('columns_number_error'))
        return False
    if titles_rows != COLS_TITLES:
        messages.add_message(request, messages.ERROR, _('columns_title_error'))
        messages.add_message(request, messages.ERROR, COLS_TITLES)
        return False
