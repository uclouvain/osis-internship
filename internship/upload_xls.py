##############################################################################
#
#    OSIS stands for Open Student Information System. It's an application
#    designed to manage the core business of higher education institutions,
#    such as universities, faculties, institutes and professional schools.
#    The core business involves the administration of students, teachers,
#    courses, programs and so on.
#
#    Copyright (C) 2015-2016 Université catholique de Louvain (http://www.uclouvain.be)
#
#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU General Public License as published by
#    the Free Software Foundation, either version 3 of the License, or
#    (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU General Public License for more details.
#
#    A copy of this license - GNU General Public License - is available
#    at the root of the source code of this program.  If not,
#    see http://www.gnu.org/licenses/.
#
##############################################################################
from django.http import HttpResponseRedirect
from django.contrib.auth.decorators import login_required
from django.core.urlresolvers import reverse
from django.utils.translation import ungettext
from openpyxl import load_workbook
from django.contrib import messages
from django.utils.translation import ugettext_lazy as _

from internship.models import Organization, OrganizationAddress, InternshipOffer, InternshipMaster
from internship.forms import OrganizationForm, InternshipOfferForm, InternshipMasterForm
from base import models as mdl

@login_required
def upload_places_file(request):
    if request.method == 'POST':
        form = OrganizationForm(request.POST, request.FILES)
        file_name = request.FILES['file']
        if file_name is not None:
            if ".xls" not in str(file_name):
                messages.add_message(request, messages.ERROR, _('file_must_be_xls'))
            else:
                __save_xls_place(request, file_name, request.user)

    return HttpResponseRedirect(reverse('internships_places'))


def __save_xls_place(request, file_name, user):
    workbook = load_workbook(file_name, read_only=True)
    worksheet = workbook.active
    form = OrganizationForm(request)
    col_reference = 0
    col_name = 1
    col_address = 2
    col_postal_code = 3
    col_city = 4
    col_country = 5
    col_url = 6

    # Iterates over the lines of the spreadsheet.
    for count, row in enumerate(worksheet.rows):
        new_score = False
        if row[col_reference].value is None \
                or row[col_reference].value == 0 \
                or not _is_registration_id(row[col_reference].value):
            continue

        reference = ""
        if row[col_reference].value < 10 :
            reference = "0"+str(row[col_reference].value)
        else :
            reference = str(row[col_reference].value)

        place = Organization.search(reference=reference)
        if place :
            organization = Organization.find_by_id(place[0].id)
        else:
            organization = Organization()

        if row[col_reference].value:
            reference = ""
            if int(row[col_reference].value) < 10 :
                reference = "0"+str(row[col_reference].value)
            else :
                reference = str(row[col_reference].value)
            organization.reference = reference
        else:
            organization.reference = None

        if row[col_name].value:
            organization.name = row[col_name].value
            organization.acronym = row[col_name].value[:14]
        else:
            organization.name = None
            organization.acronym = None

        if row[col_url].value:
            organization.website = row[col_url].value
        else:
            organization.website = ""

        organization.type = "service partner"

        organization.save()

        if place :
            organization_address = OrganizationAddress.find_by_organization(organization)
            if not organization_address:
                organization_address = OrganizationAddress()
            else :
                organization_address = organization_address[0]
        else :
            organization_address = OrganizationAddress()


        if organization:
            organization_address.label = "Addr"+organization.name[:14]
        else:
            organization_address.label = " "

        if row[col_address].value:
            organization_address.location = row[col_address].value
        else:
            organization_address.location = " "

        if row[col_postal_code].value:
            organization_address.postal_code = row[col_postal_code].value
        else:
            organization_address.postal_code = " "

        if row[col_city].value:
            organization_address.city = row[col_city].value
        else:
            organization_address.city = " "

        if row[col_country].value:
            organization_address.country = row[col_country].value
        else:
            organization_address.country = " "
        organization_address.organization = organization
        organization_address.save()

def _is_registration_id(registration_id):
    try:
        int(registration_id)
        return True
    except ValueError:
        return False


@login_required
def upload_internships_file(request):
    if request.method == 'POST':
        form = InternshipOfferForm(request.POST, request.FILES)
        file_name = request.FILES['file']
        if file_name is not None:
            if ".xls" not in str(file_name):
                messages.add_message(request, messages.ERROR, _('file_must_be_xls'))
            else:
                __save_xls_internships(request, file_name, request.user)

    return HttpResponseRedirect(reverse('internships'))


def __save_xls_internships(request, file_name, user):
    workbook = load_workbook(file_name, read_only=True)
    worksheet = workbook.active
    form = InternshipOfferForm(request)
    col_reference = 0
    col_spec = 1
    col_max_places = 4
    index = 1
    # Iterates over the lines of the spreadsheet.
    for count, row in enumerate(worksheet.rows):

        new_score = False
        if row[col_reference].value is None \
                or row[col_reference].value == 0 \
                or not _is_registration_id(row[col_reference].value):
            continue

        if row[col_spec].value is not None:
            success = 0
            check_internship = 0
            form = InternshipOfferForm(data=request.POST)

            if row[col_reference].value:
                reference = ""
                if int(row[col_reference].value) < 10 :
                    reference = "0"+str(row[col_reference].value)
                else :
                    reference = str(row[col_reference].value)
                organization = Organization.search(reference=reference)
                #internship.organization = organization[0]

            if row[col_spec].value == "CH":
                spec = "Stage en Chirurgie"
            if row[col_spec].value == "GE":
                spec = "Stage en Gériatrie"
            if row[col_spec].value == "GO":
                spec = "Stage en Gynécologie-Obstétrique"
            if row[col_spec].value == "MI":
                spec = "Stage en Médecine interne"
            if row[col_spec].value == "PE":
                spec = "Stage en Pédiatrie"
            if row[col_spec].value == "UR":
                spec = "Stage aux Urgences"

            learning_unit_year = mdl.learning_unit_year.search(title=spec)
            check_internship = InternshipOffer.find_interships_by_learning_unit_organization(spec,organization[0].name)

            if len(check_internship) != 0:
                internship = InternshipOffer.find_intership_by_id(check_internship[0].id)
            else :
                internship = InternshipOffer()

            internship.organization = organization[0]
            internship.learning_unit_year = learning_unit_year[0]
            internship.title = spec
            internship.maximum_enrollments = row[col_max_places].value
            internship.selectable = True
            
            internship.save()

@login_required
def upload_masters_file(request):
    if request.method == 'POST':
        form = InternshipOfferForm(request.POST, request.FILES)
        file_name = request.FILES['file']
        if file_name is not None:
            if ".xls" not in str(file_name):
                messages.add_message(request, messages.ERROR, _('file_must_be_xls'))
            else:
                __save_xls_masters(request, file_name, request.user)

    return HttpResponseRedirect(reverse('interships_masters'))

@login_required
def __save_xls_masters(request, file_name, user):
    workbook = load_workbook(file_name, read_only=True)
    worksheet = workbook.active
    form = InternshipMasterForm(request)

    col_reference = 2
    col_firstname = 3
    col_lastname = 4
    col_mail = 7
    col_organization_reference = 6
    col_civility = 0
    col_mastery = 1
    col_speciality = 5

    # Iterates over the lines of the spreadsheet.
    for count, row in enumerate(worksheet.rows):

        new_score = False
        if row[col_reference].value is None \
                or row[col_reference].value == 0 \
                or not _is_registration_id(row[col_reference].value):
            continue


        form = InternshipMasterForm(data=request.POST)
        if row[col_firstname].value and row[col_lastname].value:
            master = InternshipMaster.find_master_by_firstname_name(row[col_firstname].value, row[col_lastname].value)
            if len(master) == 0:
                master = InternshipMaster()

            if row[col_organization_reference].value:
                reference = ""
                if int(row[col_organization_reference].value) < 10 :
                    reference = "0"+str(row[col_organization_reference].value)
                else :
                    reference = str(row[col_organization_reference].value)
                organization = Organization.search(reference=reference)

                master.organization = organization[0]

            if row[col_firstname].value:
                master.first_name = row[col_firstname].value
            else :
                master.first_name = " "

            if row[col_lastname].value:
                master.last_name = row[col_lastname].value
            else :
                master.last_name = " "

            if row[col_reference].value:
                master.reference = row[col_reference].value
            else:
                master.reference = " "

            if row[col_civility].value:
                master.civility = row[col_civility].value
            else:
                master.civility = " "

            if row[col_mastery].value:
                master.type_mastery = row[col_mastery].value
            else:
                master.type_mastery = " "

            if row[col_speciality].value:
                master.speciality = row[col_speciality].value
            else:
                master.speciality = " "

            master.save()
