import functools
import io
from unittest import mock

import faker
import openpyxl
from django.contrib.auth.models import Permission, User
from django.core.urlresolvers import reverse
from django.test import TestCase
from openpyxl.writer.excel import save_virtual_workbook

from internship.models.organization import Organization
from internship.tests.factories.cohort import CohortFactory
from internship.utils.upload_xls import _save_xls_place as save_xls_place

fake = faker.Faker()


class XlsPlaceImportTestCase(TestCase):
    def setUp(self):
        self.cohort = CohortFactory()
        self.user = User.objects.create_user('demo',
                                             email='demo@demo.org',
                                             password='password')
        permission = Permission.objects.get(codename='is_internship_manager')
        self.user.user_permissions.add(permission)
        self.user.save()

    @classmethod
    def generate_workbook(cls):
        col_reference, col_name, col_address, \
            col_postal_code, col_city, col_country, \
            col_url = range(0, 7)

        columns = (
            (col_reference, lambda: fake.random_int(1, 500)),
            (col_name, lambda: 'Hospital {}'.format(fake.name())),
            (col_address, lambda: fake.address().replace('\n', '')),
            (col_postal_code, fake.postalcode),
            (col_city, fake.city),
            (col_country, fake.country),
            (col_url, fake.url)
        )

        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        for row in range(1, 11):
            add_cell = functools.partial(worksheet.cell, row=row)
            for column, function in columns:
                add_cell(column=column+1, value=function())

        return workbook

    def test_import_xls_from_web(self):
        self.client.force_login(self.user)

        workbook = self.generate_workbook()

        content_of_workbook = io.BytesIO(save_virtual_workbook(workbook))
        content_of_workbook.name = 'demo.xls'

        url = reverse('upload_places', kwargs={
            'cohort_id': self.cohort.id,
        })
        response = self.client.post(url, {'file': content_of_workbook}, follow=True)

        # print(response.status_code)
        self.assertRedirects(response, reverse('internships_places', kwargs={
            'cohort_id': self.cohort.id,
        }))

    def test_import_xls_from_python_api(self):

        # workbook.save('/tmp/demo.xls')
        with mock.patch('openpyxl.load_workbook') as mock_load_workbook:
            workbook = self.generate_workbook()
            mock_load_workbook.return_value = workbook

            qs = Organization.objects.all()
            self.assertEqual(qs.count(), 0)
            save_xls_place(None, self.user, cohort=self.cohort)

            self.assertTrue(mock_load_workbook.called)
            qs = Organization.objects.all()
            self.assertTrue(qs.count() > 0)
