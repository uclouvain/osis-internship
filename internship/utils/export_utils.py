##############################################################################
#
#    OSIS stands for Open Student Information System. It's an application
#    designed to manage the core business of higher education institutions,
#    such as universities, faculties, institutes and professional schools.
#    The core business involves the administration of students, teachers,
#    courses, programs and so on.
#
#    Copyright (C) 2015-2016 Université catholique de Louvain (http://www.uclouvain.be)
#
#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU General Public License as published by
#    the Free Software Foundation, either version 3 of the License, or
#    (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU General Public License for more details.
#
#    A copy of this license - GNU General Public License - is available
#    at the root of the source code of this program.  If not,
#    see http://www.gnu.org/licenses/.
#
##############################################################################
import datetime
from django.http import HttpResponse, HttpResponseRedirect
from django.core.urlresolvers import reverse
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import Color, Style, PatternFill
from django.utils.translation import ugettext_lazy as _
import re

from internship.models.organization import Organization
import internship.models as mdl_internship

HEADER = [str(_('lastname')),
          str(_('firstname')),
          str(_('noma')),
          str(_('email')),
          str(_('addresses')),
          str(_('gender')),
          str(_('birth_date')),
          str(_('mobile_phone'))]


def export_xls(cohort, organization, affections_by_specialities, file_name):
    if not affections_by_specialities:
        redirect_url = reverse('place_detail_student_affectation', kwargs={
            'cohort_id': cohort.id,
            'organization_id': organization.id,
        })
        return HttpResponseRedirect(redirect_url)

    periods = mdl_internship.period.search(cohort=cohort)

    workbook = Workbook()
    if workbook.worksheets:
        workbook.remove_sheet(workbook.worksheets[0])
    for speciality, affectations in affections_by_specialities:
        pattern = re.compile('Stage en|Intenship in', re.IGNORECASE)
        sheet_title = pattern.sub('', speciality.name.strip())[0:30]
        worksheet = workbook.create_sheet(title=sheet_title)
        worksheet.append([str(organization.name) if organization else ''])
        worksheet.append([str(speciality.name)])
        worksheet.append([str('')])
        printing_date = datetime.datetime.now()
        printing_date = printing_date.strftime("%d/%m/%Y")
        worksheet.append([str('%s: %s' % (_('file_production_date'), printing_date))])
        worksheet.append([str('')])
        worksheet.append([str(affectations[0].master) if affectations else ''])
        worksheet.append([str('')])
        worksheet.append([str('')])

        __columns_resizing(worksheet)

        row_number = 9
        for period in periods:
            worksheet.append([str(period.name), period.date_start.strftime("%d-%m-%Y"), period.date_end.strftime("%d-%m-%Y")])
            __coloring_non_editable(worksheet, row_number)
            row_number += 1
            for affectation in affectations:
                if affectation.period.name == period.name:
                    student = affectation.student
                    worksheet.append([str(student.person.last_name),
                                      str(student.person.first_name),
                                      student.registration_id,
                                      affectation.email,
                                      affectation.adress,
                                      student.person.gender,
                                      student.person.birth_date,
                                      affectation.phone_mobile])
                    row_number += 1

            worksheet.append([str('')])
            row_number += 1

    #filename_speciality = str(affectations[0].speciality.acronym).strip()
    filename = "affectation_%s_%s.xlsx" % (str(organization.reference),
                                           file_name)
    response = HttpResponse(save_virtual_workbook(workbook), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=%s' % filename
    return response


def __columns_resizing(ws):
    """
    Definition of the columns sizes
    """
    col_academic_year = ws.column_dimensions['A']
    col_academic_year.width = 18
    col_academic_year = ws.column_dimensions['B']
    col_academic_year.width = 18
    col_academic_year = ws.column_dimensions['C']
    col_academic_year.width = 18
    col_academic_year = ws.column_dimensions['D']
    col_academic_year.width = 40
    col_academic_year = ws.column_dimensions['E']
    col_academic_year.width = 40
    col_last_name = ws.column_dimensions['F']
    col_last_name.width = 10
    col_first_name = ws.column_dimensions['G']
    col_first_name.width = 20
    col_phone = ws.column_dimensions['H']
    col_phone.width = 30

def __coloring_non_editable(ws, row_number):
    """
    Coloring of the non-editable columns
    """
    style_no_modification = Style(fill=PatternFill(patternType='solid', fgColor=Color('C1C1C1')))
    column_number = 1
    while column_number < 9:
        ws.cell(row=row_number, column=column_number).style = style_no_modification

        column_number += 1
