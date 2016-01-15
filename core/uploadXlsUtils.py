##############################################################################
#
#    OSIS stands for Open Student Information System. It's an application
#    designed to manage the core business of higher education institutions,
#    such as universities, faculties, institutes and professional schools.
#    The core business involves the administration of students, teachers,
#    courses, programs and so on.
#
#    Copyright (C) 2015-2016 Université catholique de Louvain (http://www.uclouvain.be)
#
#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU General Public License as published by
#    the Free Software Foundation, either version 3 of the License, or
#    (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU General Public License for more details.
#
#    A copy of this license - GNU General Public License - is available
#    at the root of the source code of this program.  If not,
#    see http://www.gnu.org/licenses/.
#
##############################################################################
from django.http import HttpResponse
from django.http import HttpResponseRedirect
from django.contrib.auth.decorators import login_required
from django.core.urlresolvers import reverse
from django.utils.translation import ugettext_lazy as _
from django.utils.translation import ungettext

from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.cell import get_column_letter
from openpyxl import load_workbook

from core.models import AcademicCalendar, SessionExam, ExamEnrollment, LearningUnitYear, Person, AcademicYear, Student,OfferYear,LearningUnitEnrollment,OfferEnrollment
from core.forms import ScoreFileForm

from django.contrib import messages
from . import exportUtils


@login_required
def upload_scores_file(request, session_id, learning_unit_year_id, academic_year_id):
    """

    :param request:
    :return:
    """
    message_validation = ""
    if request.method == 'POST':
        form = ScoreFileForm(request.POST, request.FILES)
        if form.is_valid():
            file_name = request.FILES['file']
            if  file_name == None:
                #todo vérifier si fichier xls
                message_validation = _('No file selected')
            else:
                # if file_name.find(".xls") == -1:
                if ".xls" not in str(file_name):
                    isValid = False
                    message_validation = _('The file must be a XLS')
                else:
                    isValid = __save_xls_scores(request, file_name)
                    #todo afficher un message parlant dans l'écran si xls invalide ou problème
                    if isValid:
                        pass
                    else:
                        request.session['message_validation'] = _('Invalid file')

                messages.add_message(request, messages.INFO, '%s' % message_validation)
                return HttpResponseRedirect(reverse('online_encoding' , args=[session_id]))
        else:
            #todo traiter si pas de fichier sélectionné
            return HttpResponseRedirect(reverse('online_encoding' , args=[session_id]))


def __save_xls_scores(request, file_name):
    wb = load_workbook(file_name, read_only=True)
    ws = wb.active
    nb_row = 0
    isValid = True
    erreur_validation = ""
    data_line_number = 1
    nb_nouvelles_notes = 0
    nouvelles_notes = False
    for row in ws.rows:
        if nb_row > 0 and isValid:
            student = Student.objects.filter(registration_id=row[4].value)
            info_line = "%s %d :" % (_('Line'),data_line_number)
            if not student:
                erreur_validation += "%s %s (%s) %s!" % (info_line, _('the student'), str(row[4].value), _('does not exists'))
            else:
                academic_year = AcademicYear.objects.filter(year=int(row[0].value[:4]))
                if not academic_year :
                    erreur_validation += "%s %s (%d) %s!" % (info_line, _('the academic year'), row[0].value, _('does not exists'))
                else:
                    offer_year = OfferYear.objects.filter(academic_year=academic_year,acronym=row[3].value)
                    if  not offer_year :
                        erreur_validation += "%s %s (%d) %s-%d !" % (info_line, _('the offer year'), str(row[3].value), academic_year.year, _('does not exists'))
                    else:
                        offer_enrollment = OfferEnrollment.objects.filter(student=student,offer_year=offer_year)
                        if not offer_enrollment :
                            erreur_validation += "%s %s %s!" % (info_line, _('the offer enrollment'), _('does not exists'))
                        else:
                            learning_unit_year_lists = LearningUnitYear.objects.filter(academic_year=academic_year,acronym=row[2].value)
                            if len(learning_unit_year_lists) == 1:
                                learning_unit_year = learning_unit_year_lists[0]
                            if not learning_unit_year:
                                erreur_validation += "%s %s %s %s!" % (info_line, _('the activity'), str(row[2].value), _('does not exists'))
                            else:
                                learning_unit_enrollment = LearningUnitEnrollment.objects.filter(learning_unit_year=learning_unit_year,offer_enrollment=offer_enrollment)
                                if not learning_unit_enrollment:
                                    erreur_validation += "%s %s %s %s!" % (info_line, _('the enrollment to the activity'), str(row[2].value), _('does not exists'))
                                else:
                                    exam_enrollment = ExamEnrollment.objects.filter(learning_unit_enrollment = learning_unit_enrollment).filter(session_exam__number_session = int(row[1].value)).first()
                                    if exam_enrollment.encoding_status != 'SUBMITTED':

                                        if row[7].value is None:
                                            note = None
                                        else:
                                            note = float(row[7].value)
                                        note_valide = True
                                        if not(note is None):
                                            if note<0 or note>20:
                                                erreur_validation += "%s %s!" % (info_line, _('the score seems to be incorrect (it must be >=0 and <=20)'))
                                                note_valide = False
                                            else:
                                                if not(learning_unit_year is None) and learning_unit_year.credits < 15:
                                                    #vérification des décimales
                                                    if round(note) != note:
                                                        erreur_validation += "%s %s!" % (info_line, _('the score seems to be incorrect. When activity\'s credits < 15 decimales are not allowed!'))
                                                        note_valide = False

                                        if note_valide:
                                            if exam_enrollment.score != note:
                                                nb_nouvelles_notes = nb_nouvelles_notes + 1
                                                nouvelles_notes = True

                                            exam_enrollment.score = note

                                        if exam_enrollment.justification != row[8].value:
                                            nb_nouvelles_notes = nb_nouvelles_notes + 1
                                            nouvelles_notes = True

                                        exam_enrollment.justification = row[8].value
                                        exam_enrollment.save()

            data_line_number=data_line_number+1

        else:
            print ('else')
            #Il faut valider le fichier xls
            #Je valide les entêtes de colonnes
            #todo Il faut valider les données
            # list_header=[str(_(''))"Année académique",str(_(''))"Session",str(_(''))"Code cours",str(_(''))"Programme",str(_(''))"Noma",str(_(''))"Nom",str(_(''))"Prénom",str(_(''))"Note chiffrée",str(_(''))"Autre note",str(_(''))"Date de remise"]
            list_header = exportUtils.HEADER
            print ('else 2')
            i = 0
            for header_col in list_header:
                if str(row[i].value) != header_col:
                    isValid = False
                    break
                i = i +1
            nb_nouvelles_notes=0

        nb_row = nb_row + 1

    messages.add_message(request, messages.WARNING, erreur_validation)
    if nouvelles_notes :
        if nb_nouvelles_notes > 0:
            count = nb_nouvelles_notes
            text = ungettext(
                '%(count)d %(name)s.',
                '%(count)d %(plural_name)s.',
                count
            ) % {
                'count': nb_nouvelles_notes,
                'name': _('score injected'),
                'plural_name':  _('scores injected')
            }

            messages.add_message(request, messages.INFO, '%s' % text)
    else:
        messages.add_message(request, messages.INFO, '%s' % _('No score injected'))

    return isValid
