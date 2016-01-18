##############################################################################
#
#    OSIS stands for Open Student Information System. It's an application
#    designed to manage the core business of higher education institutions,
#    such as universities, faculties, institutes and professional schools.
#    The core business involves the administration of students, teachers,
#    courses, programs and so on.
#
#    Copyright (C) 2015-2016 Université catholique de Louvain (http://www.uclouvain.be)
#
#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU General Public License as published by
#    the Free Software Foundation, either version 3 of the License, or
#    (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU General Public License for more details.
#
#    A copy of this license - GNU General Public License - is available
#    at the root of the source code of this program.  If not,
#    see http://www.gnu.org/licenses/.
#
##############################################################################
from django.http import HttpResponse

import openpyxl
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.cell import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.styles import Fill, Color, Style, PatternFill
from openpyxl.worksheet import Worksheet, ColumnDimension, RowDimension

from core.models import AcademicCalendar, SessionExam, ExamEnrollment, LearningUnitYear, Person, AcademicYear

from django.utils.dateformat import DateFormat
from django.utils.formats import get_format
from django.utils.translation import ugettext_lazy as _

HEADER = [str(_('Academic year')),
          str(_('Session')),
          str(_('Activity code')),
          str(_('Program')),
          str(_('Registration number')),
          str(_('Last name')),
          str(_('First name')),
          str(_('Numbered score')),
          str(_('Other score')),
          str(_('End date')),
          str(_('Credits')),
          str(_('ID'))]

def export_xls(request, session_id, learning_unit_year_id, academic_year_id, isFac):

    academic_year = AcademicYear.find_academic_year(academic_year_id)
    session_exam = SessionExam.find_session(session_id)
    academic_calendar = AcademicCalendar.find_academic_calendar_by_event_type(academic_year_id,session_exam.number_session)

    wb = Workbook()
    ws = wb.active

    __columns_ajusting(ws)

# masquage de la colonne avec l'id exam enrollment

    ws.append(HEADER)

    dv = __create_data_list_for_justification(isFac)
    ws.add_data_validation(dv)

    cptr=1
    for rec_exam_enrollment in ExamEnrollment.find_exam_enrollments(session_exam.id):
        student = rec_exam_enrollment.learning_unit_enrollment.student
        o = rec_exam_enrollment.learning_unit_enrollment.offer
        person = Person.find_person(student.person.id)
        text_credits = ""
        if not(rec_exam_enrollment.learning_unit_enrollment.learning_unit_year.credits is None):
            credits = rec_exam_enrollment.learning_unit_enrollment.learning_unit_year.credits

        ws.append([str(academic_year),
                   str(session_exam.number_session),
                   session_exam.learning_unit_year.acronym,
                   o.acronym,
                   student.registration_id,
                   person.last_name,
                   person.first_name,
                   rec_exam_enrollment.score,
                   rec_exam_enrollment.justification,
                   academic_calendar.end_date.strftime('%d/%m/%Y'),
                   credits,
                   rec_exam_enrollment.id
                   ])


        cptr = cptr+1
        __coloring_non_editable(ws,cptr, rec_exam_enrollment.encoding_status,rec_exam_enrollment.score,rec_exam_enrollment.justification)


    dv.ranges.append('I2:I'+str(cptr+100))#Ajouter 100 pour si on ajoute des enregistrements

    response = HttpResponse(content=save_virtual_workbook(wb))
    response['Content-Disposition'] = 'attachment; filename=myexport.xlsx'
    response['Content-type'] = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    return response


def __columns_ajusting(ws):
    """
    ajustement des colonnes à la bonne dimension
    :param ws:
    :return:
    """
    col_academic_year = ws.column_dimensions['A']
    col_academic_year.width = 18
    col_academic_year = ws.column_dimensions['C']
    col_academic_year.width = 18
    col_academic_year = ws.column_dimensions['E']
    col_academic_year.width = 18
    col_last_name = ws.column_dimensions['F']
    col_last_name.width = 30
    col_first_name = ws.column_dimensions['G']
    col_first_name.width = 30
    col_note = ws.column_dimensions['H']
    col_note.width = 20
    col_id_exam_enrollment = ws.column_dimensions['L']
    col_id_exam_enrollment.hidden = True


def  __create_data_list_for_justification(isFac):
    """
    Création de la liste de choix pour la justification
    :return:
    """
    dv = DataValidation(type="list", formula1='%s' % ExamEnrollment.justification_label_authorized(isFac), allow_blank=True)
    dv.error = str(_('Invalid entry, not in the list of choices'))
    dv.errorTitle = str(_('Invalid entry'))

    dv.prompt = str(_('Please choose in the list'))
    dv.promptTitle = str(_('List of choices'))
    return dv


def __coloring_non_editable(ws, cptr, encoding_status, score, justification):
    """
    définition du style pour les colonnes qu'on ne doit pas modifier
    :return:
    """
    style_no_modification = Style(fill=PatternFill(patternType='solid', fgColor=Color('C1C1C1')))
    style_submitted = Style(fill=PatternFill(patternType='solid', fgColor=Color('FF9900')))

    # coloration des colonnes qu'on ne doit pas modifier
    i=1
    while i < 13:
        if i< 8 or i>9:
            ws.cell(row=cptr, column=i).style = style_no_modification
        else:
            if encoding_status == 'SUBMITTED':
                ws.cell(row=cptr, column=8).style = style_submitted
                ws.cell(row=cptr, column=9).style = style_submitted
            else:
                if not(score is None):
                    ws.cell(row=cptr, column=8).style = style_no_modification
                if not(justification is None):
                    ws.cell(row=cptr, column=9).style = style_no_modification
        i=i+1
