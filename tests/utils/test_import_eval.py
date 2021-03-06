##############################################################################
#
#    OSIS stands for Open Student Information System. It's an application
#    designed to manage the core business of higher education institutions,
#    such as universities, faculties, institutes and professional schools.
#    The core business involves the administration of students, teachers,
#    courses, programs and so on.
#
#    Copyright (C) 2015-2019 Université catholique de Louvain (http://www.uclouvain.be)
#
#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU General Public License as published by
#    the Free Software Foundation, either version 3 of the License, or
#    (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU General Public License for more details.
#
#    A copy of this license - GNU General Public License - is available
#    at the root of the source code of this program.  If not,
#    see http://www.gnu.org/licenses/.
#
##############################################################################
from unittest import mock

import openpyxl
from django.contrib.auth.models import Permission, User
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase

from base.tests.factories.student import StudentFactory
from internship.utils.importing.import_eval import import_xlsx, REGISTRATION_ID_COLUMN, PERIOD_COLUMN


class XlsImportEvalTestCase(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.user = User.objects.create_user('demo',
                                             email='demo@demo.org',
                                             password='password')
        permission = Permission.objects.get(codename='is_internship_manager')
        cls.user.user_permissions.add(permission)
        cls.user.save()
        cls.file = SimpleUploadedFile(name='test', content=b'test')
        cls.students = [StudentFactory() for _ in range(0, 10)]

    def setUp(self):
        self.client.force_login(self.user)

    def generate_workbook(cls):
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        for row, student in enumerate(cls.students):
            columns = [(REGISTRATION_ID_COLUMN, student.registration_id), (PERIOD_COLUMN, 'P1')]
            for column, value in columns:
                worksheet.cell(row=row+2, column=column+1).value = value
        return workbook

    @mock.patch('internship.utils.importing.import_eval.load_workbook')
    def test_import_eval(self, mock_workbook):
        mock_workbook.return_value = self.generate_workbook()
        evaluations = import_xlsx(self.file)
        registration_ids = [eval['registration_id'] for eval in evaluations]
        self.assertSetEqual(set(registration_ids), set([student.registration_id for student in self.students]))
