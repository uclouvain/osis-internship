##############################################################################
#
#    OSIS stands for Open Student Information System. It's an application
#    designed to manage the core business of higher education institutions,
#    such as universities, faculties, institutes and professional schools.
#    The core business involves the administration of students, teachers,
#    courses, programs and so on.
#
#    Copyright (C) 2015-2019 Université catholique de Louvain (http://www.uclouvain.be)
#
#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU General Public License as published by
#    the Free Software Foundation, either version 3 of the License, or
#    (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU General Public License for more details.
#
#    A copy of this license - GNU General Public License - is available
#    at the root of the source code of this program.  If not,
#    see http://www.gnu.org/licenses/.
#
##############################################################################
import random
from unittest import mock

import openpyxl
from django.contrib.auth.models import Permission, User
from django.core.files.uploadedfile import SimpleUploadedFile
from django.test import TestCase

from base.tests.factories.student import StudentFactory
from internship.models.internship_score import InternshipScore
from internship.tests.factories.cohort import CohortFactory
from internship.tests.factories.internship_student_information import InternshipStudentInformationFactory
from internship.tests.factories.period import PeriodFactory
from internship.tests.factories.student_affectation_stat import StudentAffectationStatFactory
from internship.utils.importing.import_scores import import_xlsx

APDS_COUNT = 15
LINE_INTERVAL = 2


class XlsImportScoresTestCase(TestCase):
    @classmethod
    def setUpTestData(cls):
        cls.cohort = CohortFactory()
        cls.period = PeriodFactory(cohort=cls.cohort)
        cls.user = User.objects.create_user('demo',
                                             email='demo@demo.org',
                                             password='password')
        permission = Permission.objects.get(codename='is_internship_manager')
        cls.user.user_permissions.add(permission)
        cls.user.save()
        cls.file = SimpleUploadedFile(name='test', content=b'test')

    def setUp(self):
        self.client.force_login(self.user)
        self.workbook = self.generate_workbook()

    @mock.patch('internship.utils.importing.import_scores.load_workbook')
    def test_import_scores(self, mock_workbook):
        mock_workbook.return_value = self.workbook
        import_xlsx(self.cohort, self.file, self.period.name)
        self.assertEqual(InternshipScore.objects.count(), 10)

    @mock.patch('internship.utils.importing.import_scores.load_workbook')
    def test_import_scores_abort_with_wrong_registration_id(self, mock_workbook):
        row_error_number = 6
        invalid_registration_id = 'invalid registration_id'
        workbook = self.workbook
        workbook.worksheets[0].cell(row=row_error_number, column=1).value = invalid_registration_id
        mock_workbook.return_value = workbook
        errors = import_xlsx(self.cohort, self.file, self.period.name)
        self.assertEqual(InternshipScore.objects.count(), 0)
        for row_error in errors['registration_error']:
            self.assertEqual(row_error[0].row, row_error_number)
            self.assertEqual(row_error[0].value, invalid_registration_id)

    @mock.patch('internship.utils.importing.import_scores.load_workbook')
    def test_import_scores_abort_with_wrong_period(self, mock_workbook):
        workbook = self.workbook
        workbook.worksheets[0].cell(row=1, column=1).value = 'invalid_period'
        mock_workbook.return_value = workbook
        errors = import_xlsx(self.cohort, self.file, self.period.name)
        self.assertEqual(errors['period_error'], 'invalid_period')
        self.assertEqual(InternshipScore.objects.count(), 0)

    def generate_workbook(self):
        students = [StudentFactory() for _ in range(0, 10)]
        [InternshipStudentInformationFactory(person=student.person, cohort=self.cohort) for student in students]
        [StudentAffectationStatFactory(student=student, period=self.period) for student in students]
        workbook = openpyxl.Workbook()
        worksheet = workbook.active
        worksheet.cell(row=1, column=1).value = 'PERIOD{}'.format(self.period.name)
        for row, student in enumerate(students):
            columns = [(0, student.registration_id), (1, '1')]
            for i in range(1, APDS_COUNT * LINE_INTERVAL, LINE_INTERVAL):
                columns.append(
                    (LINE_INTERVAL+1+i, random.choice(['A', 'B', 'C', 'D', 'E', None]))
                )
            for column, value in columns:
                worksheet.cell(row=row+6, column=column+1).value = value
        return workbook
