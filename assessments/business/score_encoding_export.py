##############################################################################
#
#    OSIS stands for Open Student Information System. It's an application
#    designed to manage the core business of higher education institutions,
#    such as universities, faculties, institutes and professional schools.
#    The core business involves the administration of students, teachers,
#    courses, programs and so on.
#
#    Copyright (C) 2015-2017 Université catholique de Louvain (http://www.uclouvain.be)
#
#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU General Public License as published by
#    the Free Software Foundation, either version 3 of the License, or
#    (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU General Public License for more details.
#
#    A copy of this license - GNU General Public License - is available
#    at the root of the source code of this program.  If not,
#    see http://www.gnu.org/licenses/.
#
##############################################################################
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import Color, Style, PatternFill, Font, colors
from django.utils.translation import ugettext_lazy as _
from django.utils import timezone

from base import models as mdl
from base.models.enums import exam_enrollment_justification_type

HEADER = [str(_('academic_year')),
          str(_('sessionn')),
          str(_('learning_unit')),
          str(_('program')),
          str(_('registration_number')),
          str(_('lastname')),
          str(_('firstname')),
          str(_('numbered_score')),
          str(_('justification')),
          str(_('end_date')),
          str(_('ID'))]

JUSTIFICATION_ALIASES = {
    exam_enrollment_justification_type.ABSENCE_JUSTIFIED : "M",
    exam_enrollment_justification_type.ABSENCE_UNJUSTIFIED : "S",
    exam_enrollment_justification_type.CHEATING : "T",
}


def export_xls(exam_enrollments):
    workbook = Workbook()
    worksheet = workbook.active

    worksheet.append([str(exam_enrollments[0].learning_unit_enrollment.learning_unit_year)])
    worksheet.append([str('Session: %s' % exam_enrollments[0].session_exam.number_session)])
    worksheet.append([str('')])
    __display_creation_date_with_message_about_state(worksheet, row_number=4)
    __display_warning_about_students_deliberated(worksheet, row_number=5)
    worksheet.append([str('')])
    __display_legends(worksheet)
    worksheet.append([str('')])
    __columns_resizing(worksheet)
    worksheet.append(HEADER)

    row_number = 11
    for exam_enroll in exam_enrollments:
        student = exam_enroll.learning_unit_enrollment.student
        offer = exam_enroll.learning_unit_enrollment.offer
        person = mdl.person.find_by_id(student.person.id)
        end_date = __get_session_exam_deadline(exam_enroll)

        score = None
        if exam_enroll.score_final is not None:
            if exam_enroll.session_exam.learning_unit_year.decimal_scores:
                score = "{0:.2f}".format(exam_enroll.score_final)
            else:
                score = "{0:.0f}".format(exam_enroll.score_final)

        justification = JUSTIFICATION_ALIASES.get(exam_enroll.justification_final, "")

        worksheet.append([str(exam_enroll.learning_unit_enrollment.learning_unit_year.academic_year),
                          str(exam_enroll.session_exam.number_session),
                          exam_enroll.session_exam.learning_unit_year.acronym,
                          offer.acronym,
                          student.registration_id,
                          person.last_name,
                          person.first_name,
                          score,
                          str(justification),
                          end_date,
                          exam_enroll.id])

        row_number += 1
        __coloring_non_editable(worksheet, row_number, score, exam_enroll.justification_final)

    lst_exam_enrollments = list(exam_enrollments)
    number_session = lst_exam_enrollments[0].session_exam.number_session
    learn_unit_acronym = lst_exam_enrollments[0].session_exam.learning_unit_year.acronym
    academic_year = lst_exam_enrollments[0].learning_unit_enrollment.learning_unit_year.academic_year

    filename = "session_%s_%s_%s.xlsx" % (str(academic_year.year), str(number_session), learn_unit_acronym)
    response = HttpResponse(save_virtual_workbook(workbook), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename=%s' % filename
    return response


def __columns_resizing(ws):
    """
    Definition of the columns sizes
    """
    col_academic_year = ws.column_dimensions['A']
    col_academic_year.width = 18
    col_academic_year = ws.column_dimensions['C']
    col_academic_year.width = 18
    col_academic_year = ws.column_dimensions['E']
    col_academic_year.width = 18
    col_last_name = ws.column_dimensions['F']
    col_last_name.width = 30
    col_first_name = ws.column_dimensions['G']
    col_first_name.width = 30
    col_note = ws.column_dimensions['H']
    col_note.width = 20
    col_note = ws.column_dimensions['I']
    col_note.width = 20
    col_id_exam_enrollment = ws.column_dimensions['K']
    # Hide the exam_enrollment_id column
    col_id_exam_enrollment.hidden = True


def __coloring_non_editable(ws, row_number, score, justification):
    """
    Coloring of the non-editable columns
    """
    pattern_fill_grey = PatternFill(patternType='solid', fgColor=Color('C1C1C1'))
    style_no_modification = Style(fill=pattern_fill_grey)
    column_number = 1
    while column_number < 12:
        if column_number < 8 or column_number > 9:
            ws.cell(row=row_number, column=column_number).style = style_no_modification
        else:
            if not(score is None and justification is None):
                ws.cell(row=row_number, column=8).style = style_no_modification
                ws.cell(row=row_number, column=9).style = style_no_modification

        column_number += 1


def __display_creation_date_with_message_about_state(ws, row_number):
    date_format = str(_('date_format'))
    printing_date = timezone.now()
    printing_date = printing_date.strftime(date_format)

    ws.cell(row=row_number, column=1).value = str('%s' % (_('warn_user_data_can_change') % printing_date))
    ws.cell(row=row_number, column=1).font = Font(color=colors.RED)


def __display_warning_about_students_deliberated(ws, row_number):
    ws.cell(row=row_number, column=1).value = str(_('students_deliberated_are_not_shown'))
    ws.cell(row=row_number, column=1).font = Font(color=colors.RED)


def __display_legends(ws):
    ws.append([
        str(_('justification')),
        str(_('justification_values_accepted') % mdl.exam_enrollment.justification_label_authorized())
    ])
    ws.append([
        str(''),
        str(_('justification_other_values') % justification_other_values())
    ])
    ws.append([
        str(_('numbered_score')),
        str(_('score_legend') % "0 - 20")
    ])


def justification_other_values():
    return "%s, %s" % (_('unjustified_absence_export_legend'),
                       _('justified_absence_export_legend'))


def __get_session_exam_deadline(exam_enroll):
    date_format = str(_('date_format'))
    deadline = None

    session_exam_deadline = mdl.exam_enrollment.get_session_exam_deadline(exam_enroll)
    if session_exam_deadline:
        deadline = session_exam_deadline.deadline_tutor_computed if session_exam_deadline.deadline_tutor_computed else\
                   session_exam_deadline.deadline

    return deadline.strftime(date_format) if deadline else "-"
