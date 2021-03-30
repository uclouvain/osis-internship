##############################################################################
#
#    OSIS stands for Open Student Information System. It's an application
#    designed to manage the core business of higher education institutions,
#    such as universities, faculties, institutes and professional schools.
#    The core business involves the administration of students, teachers,
#    courses, programs and so on.
#
#    Copyright (C) 2015-2019 Université catholique de Louvain (http://www.uclouvain.be)
#
#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU General Public License as published by
#    the Free Software Foundation, either version 3 of the License, or
#    (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU General Public License for more details.
#
#    A copy of this license - GNU General Public License - is available
#    at the root of the source code of this program.  If not,
#    see http://www.gnu.org/licenses/.
#
##############################################################################
import re

from django.db import transaction
from openpyxl import load_workbook

from base.models import student
from internship.models.internship_score import InternshipScore
from internship.models.internship_student_information import find_by_person

APDS_COUNT = 15
LINE_INTERVAL = 2
NUMBER_REGEX = r'(\d+)'


@transaction.atomic
def import_xlsx(cohort, xlsxfile, period):
    workbook = load_workbook(filename=xlsxfile, read_only=True)
    worksheet = workbook.active
    period = cohort.period_set.get(name=period)
    worksheet_period = list(worksheet.rows)[0][0].value
    errors = _search_worksheet_for_errors(cohort, period, worksheet, worksheet_period)
    if errors:
        return errors
    else:
        _process_rows_import(cohort, period, worksheet)
    xlsxfile.close()


def _process_rows_import(cohort, period, worksheet):
    for row in list(worksheet.rows)[5:worksheet.max_row]:
        _import_score(row, cohort, period)


def _search_worksheet_for_errors(cohort, period, worksheet, worksheet_period):
    errors = {}
    if not _periods_match(period, worksheet_period):
        errors.update({'period_error': worksheet_period})
    else:
        registration_error = _analyze_registration_ids(cohort, worksheet)
        if registration_error:
            errors.update({'registration_error': registration_error})
    return errors


def _analyze_registration_ids(cohort, worksheet):
    errors = []
    for row in list(worksheet.rows)[5:worksheet.max_row]:
        registration_id = row[0].value
        if registration_id is None:
            continue
        else:
            existing_student = student.find_by_registration_id(registration_id)
            if existing_student is None or not _student_is_in_cohort(existing_student, cohort):
                errors.append(row)
    return errors


def _import_score(row, cohort, period):
    scores = []
    registration_id = row[0].value
    if registration_id is None:
        return
    for i in range(1, APDS_COUNT*LINE_INTERVAL, LINE_INTERVAL):
        scores.append(row[i+LINE_INTERVAL+1].value)
    existing_student = student.find_by_registration_id(registration_id)
    internship_score, created = InternshipScore.objects.get_or_create(
        student_affectation__student=existing_student, student_affectation__period=period, validated=True
    )
    for index, score in enumerate(scores):
        internship_score.__setattr__('APD_{}'.format(index+1), score)
    internship_score.save()


def _student_is_in_cohort(student, cohort):
    return find_by_person(student.person, cohort)


def _periods_match(period, worksheet_period):
    period_numeric = re.findall(NUMBER_REGEX, period.name)
    worksheet_period_numeric = re.findall(NUMBER_REGEX, worksheet_period)
    return period_numeric[0] == worksheet_period_numeric[0] if period_numeric and worksheet_period_numeric else False
