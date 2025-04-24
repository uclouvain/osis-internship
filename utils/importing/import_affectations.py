##############################################################################
#
#    OSIS stands for Open Student Information System. It's an application
#    designed to manage the core business of higher education institutions,
#    such as universities, faculties, institutes and professional schools.
#    The core business involves the administration of students, teachers,
#    courses, programs and so on.
#
#    Copyright (C) 2015-2025 Université catholique de Louvain (http://www.uclouvain.be)
#
#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU General Public License as published by
#    the Free Software Foundation, either version 3 of the License, or
#    (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#
#    A copy of this license - GNU General Public License - is available
#    at the root of the source code of this program.  If not,
#    see http://www.gnu.org/licenses/.
#
##############################################################################
import openpyxl
from django.core.exceptions import ValidationError
from django.db.transaction import TransactionManagementError
from django.utils.translation import gettext_lazy as _

from base.models import student
from internship.models import internship_speciality, internship_student_affectation_stat, organization
from internship.models.enums.choice_type import ChoiceType
from internship.models.internship import Internship
from internship.models.internship_score import InternshipScore
from internship.models.organization import Organization

INTERNSHIP_TYPE_MANDATORY = 'Stage obligatoire'
MEDECINE_GENERALE_ACRONYM = 'MG'
MEDECINE_GENERALE_ORG_REF = 600


def import_xlsx(cohort, xlsxfile, period_instance):
    workbook = openpyxl.load_workbook(filename=xlsxfile, read_only=True)
    worksheet = workbook.active
    errors = []
    row_count = 0
    organization_mg = Organization.objects.get(cohort=cohort, reference=MEDECINE_GENERALE_ORG_REF)

    # Check if affectations already exist for the cohort's period
    existing_affectations = internship_student_affectation_stat.InternshipStudentAffectationStat.objects.filter(
        period=period_instance,
    ).exists()

    if existing_affectations:
        errors.append(_("Affectations already exist for this cohort and period. Import cancelled."))
        return errors, 0

    for index, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        # Start from the second row, index starts at 2
        row_errors = _validate_row(cohort, row, index)
        if row_errors:
            errors.extend(row_errors)
            return errors, 0

    for index, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        registration_id = row[2]
        affectation_str = row[7]
        internship_type = row[10]
        # Check if date columns exist and have values
        date_start_str = row[11] if len(row) > 11 and row[11] else None
        date_end_str = row[12] if len(row) > 12 and row[12] else None
        if registration_id and affectation_str:
            affectation_strings = affectation_str.split('/')

            # Split dates if they contain '/' to match with multiple affectations
            date_starts = date_start_str.split('/') \
                if date_start_str and '/' in date_start_str else [date_start_str] * len(affectation_strings)
            date_ends = date_end_str.split('/') \
                if date_end_str and '/' in date_end_str else [date_end_str] * len(affectation_strings)

            # Ensure we have the same number of dates as affectations
            if len(date_starts) != len(affectation_strings):
                date_starts = [date_start_str] * len(affectation_strings)
            if len(date_ends) != len(affectation_strings):
                date_ends = [date_end_str] * len(affectation_strings)
            try:
                for i, affectation_string in enumerate(affectation_strings):
                    _create_affectation(
                        cohort, period_instance, registration_id,
                        affectation_string, internship_type, index, organization_mg,
                        date_starts[i], date_ends[i]
                    )
            except ValidationError as e:
                error_msg = f"Error creating affectations: {e}"
                errors.append(error_msg)
            except TransactionManagementError as e:
                pass
            row_count += 1
    return errors, row_count


def _validate_row(cohort, row, row_index):
    errors = []
    registration_id = row[2]
    affectation_str = row[7]
    internship_type = row[10]
    # Check if date columns exist and have values
    date_start_str = row[11] if len(row) > 11 and row[11] else None
    date_end_str = row[12] if len(row) > 12 and row[12] else None

    if not registration_id or not affectation_str:
        return errors  # Skip empty rows

    # Get all affectations
    affectation_strings = affectation_str.split('/')

    # Get all dates
    date_starts = date_start_str.split('/') if date_start_str and '/' in date_start_str else [date_start_str]
    date_ends = date_end_str.split('/') if date_end_str and '/' in date_end_str else [date_end_str]

    # Validate dates if provided
    if len(date_starts) == len(date_ends):
        for i, (start, end) in enumerate(zip(date_starts, date_ends)):
            if start and end and start > end:
                errors.append(f"Row {row_index}: Start date must be earlier than end date for affectation {i+1}.")
    elif date_start_str and date_end_str:
        errors.append(
            f"Row {row_index}: Number of start dates"
            f" ({len(date_starts)}) does not match number of end dates ({len(date_ends)})."
        )

    # If number of dates doesn't match number of affectations, validate using the original strings
    if (
            len(date_starts) != len(affectation_strings) or len(date_ends) != len(affectation_strings)
    ) and date_start_str and date_end_str:
        if date_start_str > date_end_str:
            errors.append(f"Row {row_index}: Start date must be earlier than end date.")

    student_obj = student.find_by_registration_id(registration_id)
    if not student_obj:
        errors.append(f"Row {row_index}: Student with registration_id {registration_id} not found")

    affectation_strings = affectation_str.split('/')
    for affectation_string in affectation_strings:
        specialty_acronym = "".join([char for char in affectation_string if char.isalpha()])
        org_reference = "".join([char for char in affectation_string if char.isdigit()])

        specialty = internship_speciality.InternshipSpeciality.objects.filter(
            acronym=specialty_acronym, cohort=cohort
        ).first()
        if not specialty:
            errors.append(f"Row {row_index}: Specialty with acronym {specialty_acronym} not found")

        if specialty_acronym and specialty_acronym != MEDECINE_GENERALE_ACRONYM and not org_reference:
            errors.append(f"Row {row_index}: Organization reference is required for specialty {specialty_acronym}")

        if org_reference:
            organization_obj = organization.Organization.objects.filter(reference=org_reference, cohort=cohort).first()
            if not organization_obj:
                errors.append(f"Row {row_index}: Organization with reference {org_reference} not found")

    internship = None
    if internship_type == INTERNSHIP_TYPE_MANDATORY:
        internship = Internship.objects.filter(speciality=specialty, cohort=cohort).first()
    else:
        internship = Internship.objects.filter(name=internship_type, cohort=cohort).first()
    if not internship:
        errors.append(f"Row {row_index}: Internship {internship_type} not found")

    return errors


def _create_affectation(
        cohort, period_instance, registration_id, affectation_str, internship_type, row_index, organization_mg,
        date_start=None, date_end=None
):
    student_obj = student.find_by_registration_id(registration_id)
    # Process a single affectation string (already split in the calling function)
    specialty_acronym = "".join([char for char in affectation_str if char.isalpha()])
    org_reference = "".join([char for char in affectation_str if char.isdigit()])
    specialty = internship_speciality.InternshipSpeciality.objects.filter(
        acronym=specialty_acronym, cohort=cohort
    ).first()
    organization_obj = organization.Organization.objects.filter(
        reference=org_reference, cohort=cohort
    ).first() if org_reference else None

    if specialty_acronym == MEDECINE_GENERALE_ACRONYM:
        organization_obj = organization_mg

    internship = None
    if internship_type == INTERNSHIP_TYPE_MANDATORY:
        internship = Internship.objects.filter(speciality=specialty, cohort=cohort).first()
    else:
        internship = Internship.objects.filter(name=internship_type, cohort=cohort).first()

    student_affectation = internship_student_affectation_stat.InternshipStudentAffectationStat.objects.create(
        student=student_obj,
        period=period_instance,
        speciality=specialty,
        organization=organization_obj,
        internship=internship,
        cost=0,
        choice=ChoiceType.IMPOSED.value,
        date_start=date_start,
        date_end=date_end,
    )
    # create empty score along with affectation
    InternshipScore.objects.create(
        student_affectation=student_affectation
    )
