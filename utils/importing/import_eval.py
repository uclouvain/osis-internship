##############################################################################
#
#    OSIS stands for Open Student Information System. It's an application
#    designed to manage the core business of higher education institutions,
#    such as universities, faculties, institutes and professional schools.
#    The core business involves the administration of students, teachers,
#    courses, programs and so on.
#
#    Copyright (C) 2015-2019 Université catholique de Louvain (http://www.uclouvain.be)
#
#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU General Public License as published by
#    the Free Software Foundation, either version 3 of the License, or
#    (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU General Public License for more details.
#
#    A copy of this license - GNU General Public License - is available
#    at the root of the source code of this program.  If not,
#    see http://www.gnu.org/licenses/.
#
##############################################################################

from openpyxl import load_workbook


def import_xlsx(xlsxfile):
    workbook = load_workbook(filename=xlsxfile, read_only=True)
    worksheet = workbook.active
    registration_ids = [
        _get_only_digits_comprehension(str(row[6].value))
        for row in list(worksheet.rows)[1:worksheet.max_row]
    ]
    xlsxfile.close()
    return registration_ids


def _get_only_digits_comprehension(value):
    return ''.join([c for c in value if c.isdigit()])
