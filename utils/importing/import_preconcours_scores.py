##############################################################################
#
#    OSIS stands for Open Student Information System. It's an application
#    designed to manage the core business of higher education institutions,
#    such as universities, faculties, institutes and professional schools.
#    The core business involves the administration of students, teachers,
#    courses, programs and so on.
#
#    Copyright (C) 2015-2025 Université catholique de Louvain (http://www.uclouvain.be)
#
#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU General Public License as published by
#    the Free Software Foundation, either version 3 of the License, or
#    (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU General Public License for more details.
#
#    A copy of this license - GNU General Public License - is available
#    at the root of the source code of this program.  If not,
#    see http://www.gnu.org/licenses/.
#
##############################################################################

from django.db import transaction
from openpyxl import load_workbook

from base.models import student
from internship.models.internship_score import InternshipScore
from internship.models.internship_student_affectation_stat import InternshipStudentAffectationStat
from internship.models.internship_student_information import find_by_person
from osis_common.utils.models import get_object_or_none

NUMBER_REGEX = r'(\d+)'


@transaction.atomic
def import_xlsx(cohort, xlsxfile, period):
    workbook = load_workbook(filename=xlsxfile, read_only=True)
    worksheet = workbook.active
    period = cohort.period_set.get(name=period)
    worksheet_period = list(worksheet.rows)[0][0].value
    errors = _search_worksheet_for_errors(cohort, period, worksheet, worksheet_period)
    if errors:
        return errors
    else:
        _process_rows_import(cohort, period, worksheet)
    xlsxfile.close()


def _process_rows_import(cohort, period, worksheet):
    for row in list(worksheet.rows)[1:worksheet.max_row]:
        if not any(cell.value for cell in row):
            continue
        _import_score(row, cohort, period)


def _search_worksheet_for_errors(cohort, period, worksheet, worksheet_period):
    errors = {}

    registration_error = _analyze_registration_ids(cohort, worksheet)
    if registration_error:
        errors.update({'registration_error': registration_error})

    score_completeness_errors = _analyze_score_completeness(worksheet)
    if score_completeness_errors:
        errors.update({'score_completeness_errors': score_completeness_errors})

    return errors


def _analyze_score_completeness(worksheet):
    errors = []
    for row in list(worksheet.rows)[1:worksheet.max_row]:
        registration_id = row[0].value
        if registration_id is None:
            continue

        behavior_score = row[6].value
        competency_score = row[7].value

        error = _get_score_completeness_error(
            behavior_score,
            competency_score,
            errors,
            registration_id,
            row
        )
        if error:
            errors.append(error)

    return errors


def _get_score_completeness_error(behavior_score, competency_score, errors, registration_id, row):
    if behavior_score is None or competency_score is None:
        return {'registration_id': registration_id, 'row': row}
    return None


def _analyze_registration_ids(cohort, worksheet):
    errors = []
    for row in list(worksheet.rows)[1:worksheet.max_row]:
        registration_id = row[0].value
        if registration_id is None:
            continue
        else:
            error = _get_registration_id_errors(cohort, errors, registration_id, row)
            if error:
                errors.append(error)
    return errors


def _get_registration_id_errors(cohort, errors, registration_id, row):
    existing_student = student.find_by_registration_id(registration_id)
    if existing_student is None or not _student_is_in_cohort(existing_student, cohort):
        return row

def _import_score(row, cohort, period):
    registration_id = row[0].value

    if registration_id is None:
        return

    speciality_acronym = row[4].value

    existing_student = student.find_by_registration_id(registration_id)
    student_affectation = get_object_or_none(
        InternshipStudentAffectationStat,
        student=existing_student,
        period=period,
        speciality__acronym=speciality_acronym,
    )

    if student_affectation:
        internship_score, created = InternshipScore.objects.get_or_create(
            student_affectation=student_affectation,
            defaults={'validated': True}
        )

        try:
            behavior_score = row[6].value
            competency_score = row[7].value

            if all(score is not None for score in [behavior_score, competency_score]):
                internship_score.behavior_score = behavior_score
                internship_score.competency_score = competency_score
                internship_score.validated = True
                internship_score.save()

        except (ValueError, TypeError):
            pass


def _student_is_in_cohort(student, cohort):
    return find_by_person(student.person, cohort)
