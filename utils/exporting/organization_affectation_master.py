##############################################################################
#
#    OSIS stands for Open Student Information System. It's an application
#    designed to manage the core business of higher education institutions,
#    such as universities, faculties, institutes and professional schools.
#    The core business involves the administration of students, teachers,
#    courses, programs and so on.
#
#    Copyright (C) 2015-2017 Université catholique de Louvain
#
#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU General Public License as published by
#    the Free Software Foundation, either version 3 of the License, or
#    (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU General Public License for more details.
#
#    A copy of this license - GNU General Public License - is available
#    at the root of the source code of this program.  If not,
#    see http://www.gnu.org/licenses/.
#
##############################################################################
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from django.utils.translation import ugettext_lazy as _
from django.utils import timezone
from internship import models
from internship.utils.exporting.spreadsheet import columns_resizing, add_row, coloring_non_editable_line


def export_master_xls(cohort, organization, affections_by_specialities):
    periods = models.period.search(cohort=cohort)

    workbook = Workbook()
    if workbook.worksheets:
        workbook.remove_sheet(workbook.worksheets[0])

    create_worksheets(workbook, organization, periods, affections_by_specialities)

    return save_virtual_workbook(workbook)


def create_worksheets(workbook, organization, periods, affections_by_specialities):
    for specialty, affectations in affections_by_specialities:
        sheet_title = specialty.name.strip()[0:30]
        worksheet = workbook.create_sheet(title=sheet_title)

        _add_header(worksheet, organization, specialty, affectations)
        columns_resizing(worksheet, {'A': 18, 'B': 18, 'C': 18, 'D': 40, 'E': 40, 'F': 10, 'G': 20, 'H': 30})
        _add_periods(worksheet, periods, affectations)


def _add_header(worksheet, organization, specialty, affectations):
    add_row(worksheet, [str(organization.name) if organization else ''])
    add_row(worksheet, [str(specialty.name)])
    add_row(worksheet)
    date_format = str(_('Date format'))
    printing_date = timezone.now()
    printing_date = printing_date.strftime(date_format)
    add_row(worksheet, [str('%s: %s' % (_('File production date'), printing_date))])
    add_row(worksheet)
    add_row(worksheet, [str(affectations[0].master) if affectations else ''])
    add_row(worksheet)
    add_row(worksheet)


def _add_periods(worksheet, periods, affectations):
    for period in periods:
        add_row(worksheet, [str(period.name),
                            period.date_start.strftime("%d-%m-%Y"),
                            period.date_end.strftime("%d-%m-%Y")])
        coloring_non_editable_line(worksheet, worksheet.max_row, 9)
        _add_affectations(worksheet, period, affectations)
        add_row(worksheet)


def _add_affectations(worksheet, period, affectations):
    for affectation in affectations:
        if affectation.period.name == period.name:
            student = affectation.student
            add_row(worksheet, [str(student.person.last_name),
                                str(student.person.first_name),
                                student.registration_id,
                                affectation.email,
                                affectation.adress,
                                student.person.gender,
                                student.person.birth_date,
                                affectation.phone_mobile])
