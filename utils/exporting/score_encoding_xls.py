##############################################################################
#
#    OSIS stands for Open Student Information System. It's an application
#    designed to manage the core business of higher education institutions,
#    such as universities, faculties, institutes and professional schools.
#    The core business involves the administration of students, teachers,
#    courses, programs and so on.
#
#    Copyright (C) 2015-2019 Université catholique de Louvain
#
#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU General Public License as published by
#    the Free Software Foundation, either version 3 of the License, or
#    (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU General Public License for more details.
#
#    A copy of this license - GNU General Public License - is available
#    at the root of the source code of this program.  If not,
#    see http://www.gnu.org/licenses/.
#
##############################################################################
from django.utils.translation import gettext as _
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from internship import models
from internship.models.period import get_effective_periods
from internship.templatetags.dictionary import is_edited
from internship.templatetags.list import get_period_score_tuple
from internship.utils.exporting.spreadsheet import columns_resizing, add_row
from osis_common.document.xls_build import save_virtual_workbook

LAST_COLUMN = 50
PERIOD_COLUMN_WIDTH = 7


def _get_columns_width():
    columns_width = {'A': 32, 'B': 16, 'C': 11}
    columns_width.update({
        get_column_letter(i): PERIOD_COLUMN_WIDTH for i in range(4, LAST_COLUMN)
    })
    return columns_width


def export_xls_with_scores(cohort, periods, students, internships):
    workbook = Workbook()
    worksheet = workbook.active
    _add_header(cohort, periods, worksheet)
    _make_complete_list(periods, students, worksheet)
    _make_internship_sheets(internships, periods, students, workbook)
    return save_virtual_workbook(workbook)


def _make_internship_sheets(internships, periods, students, workbook):
    for internship in sorted(internships):
        workbook.create_sheet(internship)
        worksheet = workbook[internship]
        _add_sheet_header(worksheet)
        _add_sheet_content(internship, periods, students, worksheet)


def _add_sheet_content(internship, periods, students, worksheet):
    for student in students:
        if student.registration_id:
            columns = []
            columns.append(student.person.last_name.upper())
            columns.append(student.person.first_name)
            columns.append(student.registration_id)
            _complete_student_row_by_internship(columns, internship, periods, student)
            add_row(worksheet, columns)


def _complete_student_row_by_internship(columns, internship, periods, student):
    for period in periods:
        if period.name in student.specialties.keys():
            if student.specialties[period.name][0]['acronym'] == internship:
                _append_row_data(columns, period, student)


def _append_row_data(columns, period, student):
    if period.name in student.organizations.keys():
        if len(student.specialties[period.name]) > 1:
            multiple_specialties = '/'.join([sp['reference'] for sp in student.organizations[period.name]])
            columns.append(multiple_specialties)
        else:
            columns.append(student.organizations[period.name][0]['reference'])
    else:
        columns.append('')
    if period.name in student.periods_scores.keys():
        columns.append(
            _retrieve_score(
                student.periods_scores[period.name]
            )
        )
    else:
        # default to zero when grades have not been submitted and period is completed
        columns.append(0 if period.is_past else '')


def _retrieve_score(period_score):
    if len(period_score) > 1:
        return sum(period_score) / len(period_score)
    else:
        period_score = period_score[0]
    if _is_dict_with_key(period_score, 'edited'):
        if _is_dict_with_key(period_score['edited'][0], 'excused'):
            return period_score['edited'][0]['excused'] or ''
        return period_score['edited'][0]['score'] or ''
    else:
        return period_score


def _is_dict_with_key(obj, key):
    return isinstance(obj, dict) and key in obj


def _make_complete_list(periods, students, worksheet):
    columns_resizing(worksheet, _get_columns_width())
    all_periods_count = get_effective_periods(periods.first().cohort.pk).count()
    for student in students:
        if student.registration_id:
            columns = []
            columns.append(student.person.last_name.upper())
            columns.append(student.person.first_name)
            columns.append(student.registration_id)
            _complete_student_row_for_all_internships(columns, periods, student)
            if periods.count() == all_periods_count:
                _append_evolution_score(columns, student.evolution_score)
                columns.append(student.evolution_score_reason)
            columns.append(_("Yes") if student.fulfill_condition else _("No"))
            add_row(worksheet, columns)


def _append_evolution_score(columns, score):
    columns.append(score['edited']['score'] if is_edited(score) else score)
    columns.append(score['computed'] if is_edited(score) else score)
    columns.append(score['edited']['score'] if is_edited(score) else '')


def _complete_student_row_for_all_internships(columns, periods, student):
    for period in periods:
        if period.name in student.specialties.keys():
            if len(student.specialties[period.name]) > 1:
                multiple_specialties = '/'.join([sp['acronym'] for sp in student.specialties[period.name]])
                columns.append(multiple_specialties)
            else:
                columns.append(student.specialties[period.name][0]['acronym'])
        else:
            columns.append('')
            
        if period.is_preconcours:
            _append_preconcours_data(columns, period, student)
        else:
            _append_row_data(columns, period, student)


def _append_preconcours_data(columns, period, student):
    period_scores = get_period_score_tuple(student.scores, period.name)
    if period_scores:
        if isinstance(period_scores, dict):
            period_scores = [period_scores]
        behavior = 0
        competency = 0
        calculated_global_score = 0
        for period_score in period_scores:
            behavior += float(period_score.get('behavior', '')) / len(period_scores)
            competency += float(period_score.get('competency', '')) / len(period_scores)
            calculated_global_score += float(period_score.get('calculated_global_score', '')) / len(period_scores)
        columns.extend([behavior, competency, calculated_global_score])
    else:
        columns.extend(['', '', ''])


def _add_sheet_header(worksheet):
    column_titles = [_("Name"), _("First name"), _("NOMA"), _("Hospital"), _("Grade")]
    add_row(worksheet, column_titles)
    cells = worksheet.iter_rows(min_row=1, max_row=1, min_col=1, max_col=27)
    for col in cells:
        for cell in col:
            cell.font = Font(bold=True)


def _add_header(cohort, periods, worksheet):
    column_titles = [_("Name"), _("First name"), _("NOMA")]
    all_periods_count = get_effective_periods(periods.first().cohort.pk).count()
    for period in periods:
        if period.is_preconcours:
            column_titles.extend([
                period.name,
                f"{period.name} - {_('Behavior')}",
                f"{period.name} - {_('Competency')}",
                f"{period.name} - {_('Global')}"
            ])
        else:
            column_titles.append(period.name)
            column_titles.append("{}+".format(period.name))
            column_titles.append("{}-Score".format(period.name))
    if periods.count() == all_periods_count:
        column_titles.append(_("Evolution"))
        column_titles.append(_("Evolution computed"))
        column_titles.append(_("Evolution edited"))
        column_titles.append(_("Reason"))
    column_titles.append(_("EPA Validation"))
    add_row(worksheet, column_titles)
    cells = worksheet.iter_rows(min_row=1, max_row=1, min_col=1, max_col=40)
    for col in cells:
        for cell in col:
            cell.font = Font(bold=True)


def _add_students(cohort, worksheet):
    students_stat = models.internship_student_affectation_stat.find_by_cohort(cohort)
    previous_student = None
    columns = []
    for student_stat in students_stat:
        # Line breaking condition.
        if student_stat.student.registration_id != previous_student:
            if len(columns) > 0:
                add_row(worksheet, columns)
                columns = []
            columns.append(student_stat.student.person.last_name.upper())
            columns.append(student_stat.student.person.first_name)
            columns.append(student_stat.student.registration_id)
            previous_student = student_stat.student.registration_id
        _add_period(columns, student_stat)
    add_row(worksheet, columns)


def _add_period(columns, student_stat):
    _deal_with_empty_periods(columns, student_stat)
    columns.append(student_stat.speciality.acronym_with_sequence())
    columns.append("{}{}".format(student_stat.speciality.acronym, student_stat.organization.reference))


def _deal_with_empty_periods(columns, student_stat):
    period_position = (student_stat.period.number() * 2) + 1  # Consider the first columns and 2 columns per period.
    if len(columns) < period_position:
        diff = period_position - len(columns)
        while diff > 0:  # Add empty columns when there is no data to the periods before the current one.
            columns.append("")
            columns.append("")
            diff -= 2  # Subtracts the number of columns appended within the while loop.
