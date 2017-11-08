##############################################################################
#
#    OSIS stands for Open Student Information System. It's an application
#    designed to manage the core business of higher education institutions,
#    such as universities, faculties, institutes and professional schools.
#    The core business involves the administration of students, teachers,
#    courses, programs and so on.
#
#    Copyright (C) 2015-2017 Université catholique de Louvain
#
#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU General Public License as published by
#    the Free Software Foundation, either version 3 of the License, or
#    (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU General Public License for more details.
#
#    A copy of this license - GNU General Public License - is available
#    at the root of the source code of this program.  If not,
#    see http://www.gnu.org/licenses/.
#
##############################################################################
import re
from openpyxl import Workbook
from openpyxl.writer.excel import save_virtual_workbook
from openpyxl.styles import Color, Style, PatternFill
from django.utils.translation import ugettext_lazy as _
from django.utils import timezone
from internship import models

MAX_COLUMN_NUMBER = 9
COLUMN_WIDTHS = {'A': 18, 'B': 18, 'C': 18, 'D': 40, 'E': 40, 'F': 10, 'G': 20, 'H': 30}


def export_xls(cohort, organization, affections_by_specialities):
    periods = models.period.search(cohort=cohort)

    workbook = Workbook()
    if workbook.worksheets:
        workbook.remove_sheet(workbook.worksheets[0])

    create_worksheets(workbook, organization, periods, affections_by_specialities)

    return save_virtual_workbook(workbook)


def create_worksheets(workbook, organization, periods, affections_by_specialities):
    for specialty, affectations in affections_by_specialities:
        pattern = re.compile('Stage en|Intenship in', re.IGNORECASE)
        sheet_title = pattern.sub('', specialty.name.strip())[0:30]
        worksheet = workbook.create_sheet(title=sheet_title)

        _add_header(worksheet, organization, specialty, affectations)
        _columns_resizing(worksheet)
        _add_periods(worksheet, periods, affectations)


def _add_header(worksheet, organization, specialty, affectations):
    _add_row(worksheet, [str(organization.name) if organization else ''])
    _add_row(worksheet, [str(specialty.name)])
    _add_row(worksheet)
    date_format = str(_('date_format'))
    printing_date = timezone.now()
    printing_date = printing_date.strftime(date_format)
    _add_row(worksheet, [str('%s: %s' % (_('file_production_date'), printing_date))])
    _add_row(worksheet)
    _add_row(worksheet, [str(affectations[0].master) if affectations else ''])
    _add_row(worksheet)
    _add_row(worksheet)


def _add_periods(worksheet, periods, affectations):
    for period in periods:
        _add_row(worksheet, [str(period.name),
                             period.date_start.strftime("%d-%m-%Y"),
                             period.date_end.strftime("%d-%m-%Y")])
        _coloring_non_editable_line(worksheet, worksheet.max_row)
        _add_affectations(worksheet, period, affectations)
        _add_row(worksheet)


def _add_affectations(worksheet, period, affectations):
    for affectation in affectations:
        if affectation.period.name == period.name:
            student = affectation.student
            _add_row(worksheet, [str(student.person.last_name),
                                 str(student.person.first_name),
                                 student.registration_id,
                                 affectation.email,
                                 affectation.adress,
                                 student.person.gender,
                                 student.person.birth_date,
                                 affectation.phone_mobile])


def _add_row(worksheet, content=None):
    if content:
        worksheet.append(content)
    else:
        worksheet.append([str('')])


def _columns_resizing(worksheet):
    for key in COLUMN_WIDTHS.keys():
        worksheet.column_dimensions[key].width = COLUMN_WIDTHS.get(key)


def _coloring_non_editable_line(worksheet, row_number):
    style_readonly = Style(fill=PatternFill(patternType='solid', fgColor=Color('C1C1C1')))
    for column_number in range(1, MAX_COLUMN_NUMBER):
        worksheet.cell(row=row_number, column=column_number).style = style_readonly
