##############################################################################
#
#    OSIS stands for Open Student Information System. It's an application
#    designed to manage the core business of higher education institutions,
#    such as universities, faculties, institutes and professional schools.
#    The core business involves the administration of students, teachers,
#    courses, programs and so on.
#
#    Copyright (C) 2015-2019 Université catholique de Louvain
#
#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU General Public License as published by
#    the Free Software Foundation, either version 3 of the License, or
#    (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU General Public License for more details.
#
#    A copy of this license - GNU General Public License - is available
#    at the root of the source code of this program.  If not,
#    see http://www.gnu.org/licenses/.
#
##############################################################################
from openpyxl.styles import Color, Style, PatternFill


def columns_resizing(worksheet, column_widths):
    for key in column_widths.keys():
        worksheet.column_dimensions[key].width = column_widths.get(key)


def add_row(worksheet, content=None):
    if content:
        worksheet.append(content)
    else:
        worksheet.append([str('')])


def coloring_non_editable_line(worksheet, row_number, max_column_number):
    style_readonly = Style(fill=PatternFill(patternType='solid', fgColor=Color('C1C1C1')))
    for column_number in range(1, max_column_number):
        worksheet.cell(row=row_number, column=column_number).style = style_readonly
