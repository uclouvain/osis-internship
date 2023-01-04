##############################################################################
#
#    OSIS stands for Open Student Information System. It's an application
#    designed to manage the core business of higher education institutions,
#    such as universities, faculties, institutes and professional schools.
#    The core business involves the administration of students, teachers,
#    courses, programs and so on.
#
#    Copyright (C) 2015-2019 Université catholique de Louvain
#
#    This program is free software: you can redistribute it and/or modify
#    it under the terms of the GNU General Public License as published by
#    the Free Software Foundation, either version 3 of the License, or
#    (at your option) any later version.
#
#    This program is distributed in the hope that it will be useful,
#    but WITHOUT ANY WARRANTY; without even the implied warranty of
#    MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#    GNU General Public License for more details.
#
#    A copy of this license - GNU General Public License - is available
#    at the root of the source code of this program.  If not,
#    see http://www.gnu.org/licenses/.
#
##############################################################################
from django.utils.translation import gettext as _
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter
from openpyxl.writer.excel import save_virtual_workbook

from internship.utils.exporting.spreadsheet import add_row

DEFAULT_SIZE = 13
FONT_SIZE = 10


def export_xls_with_places_evaluations(cohort, evaluations):
    workbook = Workbook()
    worksheet = workbook.active
    _add_header(worksheet, cohort)
    for evaluation in evaluations:
        _add_evaluation_row(worksheet, cohort, evaluation)
    _adjust_col_width(worksheet)
    return save_virtual_workbook(workbook)


def _add_header(worksheet, cohort):
    column_titles = [_("NOMA"), _("Gender"), _("Birth Date"), _('Place'), _('Specialty'), _('Period')]
    column_titles.extend([f"{item.order} - {item.statement}" for item in cohort.placeevaluationitem_set.all()])
    add_row(worksheet, column_titles)
    _set_row_font_style(worksheet, bold=True)


def _add_evaluation_row(worksheet, cohort, evaluation):
    row = [
        evaluation.affectation.student.registration_id,
        evaluation.affectation.student.person.gender,
        evaluation.affectation.student.person.birth_date,
        evaluation.affectation.organization.name,
        evaluation.affectation.speciality.name,
        evaluation.affectation.period.name,
    ]
    for item in cohort.placeevaluationitem_set.all():
        if str(item.uuid) in evaluation.evaluation.keys():
            row.append(evaluation.evaluation[str(item.uuid)]["response"])
        else:
            row.append('')
    add_row(worksheet, row)
    _set_row_font_style(worksheet, bold=False)


def _set_row_font_style(worksheet, bold):
    last_filled_row = len(worksheet['A'])
    cells = worksheet.iter_rows(min_row=last_filled_row, max_row=last_filled_row, min_col=1, max_col=40)
    for col in cells:
        for cell in col:
            cell.font = Font(name='Arial', size=10, bold=bold)


def _adjust_col_width(worksheet):
    for column_cells in worksheet.columns:
        length = max(len(str(cell.value or "")) for cell in column_cells)
        worksheet.column_dimensions[get_column_letter(column_cells[0].column)].width = length * FONT_SIZE/DEFAULT_SIZE
